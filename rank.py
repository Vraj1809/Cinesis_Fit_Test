"""
Cinesis Good Fit Test — Part B: Filter and rank loads dynamically.

Pipeline:
  1. Load the load board dynamically from the target .xlsx file.
  2. Hard-filter by data completeness, equipment, and strict weight bounds.
  3. Compute effective rate/mile over 3 haversine legs.
  4. Drop loads below the driver's minimum effective rate.
  5. Rank remaining descending, print top 3.
"""

import logging
from math import radians, sin, cos, asin, sqrt
from dataclasses import dataclass
from typing import Optional
from pathlib import Path
import openpyxl

from extract import get_driver_profile, DriverProfile

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Distance
# ---------------------------------------------------------------------------

def haversine_miles(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    R = 3958.7613  # Earth radius in statute miles
    dlat = radians(lat2 - lat1)
    dlon = radians(lon2 - lon1)
    a = (sin(dlat / 2) ** 2
         + cos(radians(lat1)) * cos(radians(lat2)) * sin(dlon / 2) ** 2)
    return 2 * R * asin(sqrt(a))


# ---------------------------------------------------------------------------
# Loads board dynamic ingestion
# ---------------------------------------------------------------------------

@dataclass
class Load:
    load_id: str
    origin: str
    origin_lat: float
    origin_lon: float
    destination: Optional[str]
    dest_lat: Optional[float]
    dest_lon: Optional[float]
    trailer: str
    weight: Optional[float]
    price: Optional[float]


_MISSING_STRINGS = frozenset({"missing", "n/a", "", "none"})

def _parse_optional_float(val) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in _MISSING_STRINGS:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def _parse_optional_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in _MISSING_STRINGS:
        return None
    return s

def load_board_from_xlsx(path: str) -> list[Load]:
    """Reads the load board dynamically from the spreadsheet."""
    logger.info(f"Loading board from {path}")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(
            f"\n\n[CRITICAL] Excel file not found at: {path}\n"
            f"Please ensure you have downloaded the 'cinesis_good_fit_test_clean.xlsx' "
            f"file and placed it in the exact same directory as this script."
        )
        
    # Check if the 'Loads' sheet exists to prevent a raw KeyError
    if "Loads" not in wb.sheetnames:
        raise ValueError(
            f"\n\n[ERROR] Missing 'Loads' sheet in the provided Excel file. "
            f"Found sheets: {wb.sheetnames}. Please ensure the tab is named correctly."
        )
        
    ws = wb["Loads"]
    loads = []
    
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[1:]:  # skip header
        if not row or row[0] is None:
            continue
        loads.append(Load(
            load_id=str(row[0]),
            origin=str(row[1]),
            origin_lat=float(row[2]),
            origin_lon=float(row[3]),
            destination=_parse_optional_str(row[4]),
            dest_lat=_parse_optional_float(row[5]),
            dest_lon=_parse_optional_float(row[6]),
            trailer=str(row[7]),
            weight=_parse_optional_float(row[8]),
            price=_parse_optional_float(row[9]),
        ))
    return loads


# ---------------------------------------------------------------------------
# Evaluation
# ---------------------------------------------------------------------------

@dataclass
class Result:
    load_id: str
    eligible: bool
    reason: str
    effective_rate: Optional[float] = None
    dh_to_origin: Optional[float] = None
    loaded_miles: Optional[float] = None
    dh_home: Optional[float] = None
    total_miles: Optional[float] = None


def evaluate(load: Load, d: DriverProfile) -> Result:
    # Gate 1 — Data completeness
    if load.price is None:
        return Result(load.load_id, False,
                      "EXCLUDED — missing price: effective rate is uncomputable")
    if load.destination is None or load.dest_lat is None or load.dest_lon is None:
        return Result(load.load_id, False,
                      "EXCLUDED — missing destination: loaded miles & deadhead-home are uncomputable")

    # Gate 2 — Equipment match
    if load.trailer.lower() not in d.equipment_types:
        return Result(load.load_id, False,
                      f"INELIGIBLE — trailer '{load.trailer}' not in driver equipment "
                      f"{[e.title() for e in d.equipment_types]}")

    # Gate 3 — Weight bounds and safety
    if load.weight is None:
        return Result(load.load_id, False,
                      "EXCLUDED — weight unknown: cannot verify equipment compatibility; "
                      "requires manual confirmation before dispatch")
    if load.weight > d.weight_capacity_lb:
        return Result(load.load_id, False,
                      f"INELIGIBLE — load weight {load.weight:,.0f} lb exceeds "
                      f"driver capacity {d.weight_capacity_lb:,.0f} lb")

    # Compute legs
    dh_to = haversine_miles(d.current_lat, d.current_lon, load.origin_lat, load.origin_lon)
    loaded = haversine_miles(load.origin_lat, load.origin_lon, load.dest_lat, load.dest_lon)
    dh_home = haversine_miles(load.dest_lat, load.dest_lon, d.home_lat, d.home_lon)
    total = dh_to + loaded + dh_home

    if total <= 0:
        return Result(load.load_id, False, "INELIGIBLE — zero total miles (origin = destination = home?)")

    eff = load.price / total

    # Gate 4 — Effective rate floor
    if eff < d.min_rate_per_mile:
        return Result(load.load_id, False,
                      f"INELIGIBLE — effective ${eff:.3f}/mi is below driver minimum "
                      f"${d.min_rate_per_mile:.2f}/mi",
                      eff, dh_to, loaded, dh_home, total)

    return Result(load.load_id, True, "ELIGIBLE", eff, dh_to, loaded, dh_home, total)


def rank_loads(loads: list[Load], driver: DriverProfile) -> tuple[list[Result], list[Result]]:
    results = [evaluate(l, driver) for l in loads]
    eligible = sorted(
        [r for r in results if r.eligible],
        key=lambda r: r.effective_rate or 0.0,
        reverse=True,
    )
    return results, eligible


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    
    # 1. Provide the exact transcript required for the assessment run
    TRANSCRIPT_FILE = Path(__file__).parent / "cinesis_good_fit_test_clean.xlsx - Sample Conversation.csv"
    
    # Fallback to the known string if the CSV isn't present
    transcript_text = """
    Driver: Like, tell me a little bit about how y'all can assist me.
    Dispatch: I think you're based out in San Antonio. Is that correct?
    Driver: Yes, that's correct. I'm usually in that area, but I'm in Dallas.
    Dispatch: So for example, I've got about 30 loads that need to be moved on Tuesday...
    Dispatch: I've got one coming out of San Antonio to Huntsville... paying $1,100.
    Driver: Do y'all deal with hotshots too, like flatbeds or goosenecks?
    Dispatch: Yes, we work with hotshots.
    Driver: I might only run two or three days a week, but I still make good money—like $2,300 to $2,500 going Corpus to Odessa, or $1,800 to $2,000 from San Antonio to Midland. I run a hotshot gooseneck trailer.
    Driver: As long as it's above $2 per mile, I'll consider it.
    """
    
    try:
        with open(TRANSCRIPT_FILE, 'r') as f:
            transcript_text = f.read()
    except FileNotFoundError:
        logger.warning(
            "\n*** TRANSCRIPT CSV NOT FOUND ***\n"
            "Please ensure the assessment CSV files are in the same directory if you wish to test file ingestion.\n"
            "Falling back to the embedded sample transcript string to continue execution...\n"
        )

    # 2. Extract Profile
    driver = get_driver_profile(transcript_text)

    # 3. Ingest Load Board
    LOADS_PATH = Path(__file__).parent / "cinesis_good_fit_test_clean.xlsx"
    loads = load_board_from_xlsx(str(LOADS_PATH))

    # 4. Rank
    all_results, eligible = rank_loads(loads, driver)

    # 5. Output
    print("\n" + "=" * 72)
    print(f"DRIVER:  {driver.current_location} (current)  |  "
          f"{driver.home_base} (home base)")
    print(f"         Equipment: {', '.join(e.title() for e in driver.equipment_types)}"
          f"  |  Capacity: {driver.weight_capacity_lb:,.0f} lb"
          f"  |  Min rate: ${driver.min_rate_per_mile:.2f}/mi")
    print("=" * 72)

    print("\nALL LOADS — evaluation")
    print("-" * 72)
    for r in all_results:
        tag = "✓" if r.eligible else "✗"
        line = f"  {tag}  {r.load_id}  {r.reason}"
        if r.effective_rate is not None and r.total_miles is not None:
            line += (f"\n        legs: {r.dh_to_origin:.1f} dh-to + "
                     f"{r.loaded_miles:.1f} loaded + "
                     f"{r.dh_home:.1f} dh-home = {r.total_miles:.1f} mi  |  "
                     f"${r.effective_rate:.3f}/mi")
        print(line)

    print("\n" + "=" * 72)
    print("TOP 3 ELIGIBLE LOADS  (ranked by effective rate/mile)")
    print("=" * 72)
    for rank, r in enumerate(eligible[:3], 1):
        print(f"  {rank}.  {r.load_id}   ${r.effective_rate:.3f}/mi"
              f"   ({r.dh_to_origin:.1f} + {r.loaded_miles:.1f} + "
              f"{r.dh_home:.1f} = {r.total_miles:.1f} mi total)")

    if len(eligible) < 3:
        print(f"\n  NOTE: only {len(eligible)} load(s) passed all filters.")